"""Connectors for published resource totals covering a whole county or state.

Helios estimates how much power and water a *site* will draw. Those estimates
are inferences from acreage, and on their own they have no scale: 40 MW is
either negligible or alarming depending on what the surrounding area already
uses. These connectors supply the denominator, and it is a measured one.

Two sources, one shape
----------------------
Both publish a single bulk table rather than a queryable API, so the shared
:class:`BulkAreaTotalsConnector` does the work and each subclass only has to say
where the file is and how to read a row out of it.

* **USGS** county-level water use, per county, in million gallons per day.
* **EIA** retail electricity sales, per state, in megawatt-hours per year.

The granularity differs and that difference is not cosmetic. Water is published
per county; electricity is published per state and no public source breaks it to
county nationally. A state figure is a much weaker denominator for a metro-scale
region, and every row records its ``area_kind`` so the UI can say which it is
rather than leaving a reader to assume they match.

Both totals are ``reported``. Helios did not derive them, and it must not
present them with the same weight as its own estimates, nor sum the two.
"""

from __future__ import annotations

import csv
import io
import zipfile
from abc import abstractmethod
from typing import Any

from helios_common.hashing import short_hash
from helios_common.logging import get_logger
from helios_common.time import utcnow
from helios_common.vocabulary import AssertionClass, ExtractionMethod
from helios_connectors.base import BaseConnector
from helios_connectors.types import (
    DateRange,
    DiscoveryResult,
    ExtractedField,
    FetchResult,
    HealthCheckResult,
    NormalizationResult,
    NormalizedRecord,
    ParsedDocument,
    ParseResult,
    RawDocument,
    SourceItem,
)

logger = get_logger(__name__)

__all__ = [
    "BulkAreaTotalsConnector",
    "EiaStateElectricityConnector",
    "UsgsCountyWaterConnector",
]


class BulkAreaTotalsConnector(BaseConnector):
    """Fetches one bulk table and emits ``area_total`` records from it."""

    document_type = "area_totals_table"

    @property
    @abstractmethod
    def download_url(self) -> str:
        """The single file this connector retrieves."""

    @property
    @abstractmethod
    def reference_year(self) -> int:
        """The year the published figures describe."""

    @abstractmethod
    def rows_from(self, payload: bytes) -> list[dict[str, Any]]:
        """Turn the raw file into flat dicts, one per area/metric measurement."""

    def discover(self, date_range: DateRange) -> DiscoveryResult:
        """Return the single file this connector retrieves.

        Args:
            date_range: Unused; these are periodic full publications.

        Returns:
            A discovery result with exactly one item.
        """
        del date_range
        meta = self.get_metadata()
        return DiscoveryResult(
            items=[
                SourceItem(
                    source_native_id=f"{meta.slug}:{self.reference_year}",
                    url=self.download_url,
                    title=f"{meta.name} ({self.reference_year})",
                    document_type=self.document_type,
                    hints={"reference_year": self.reference_year},
                )
            ]
        )

    def fetch(self, item: SourceItem) -> FetchResult:
        """Download the bulk file whole.

        Args:
            item: The discovered item.

        Returns:
            The raw document, or an error.
        """
        try:
            response = self.http.get(item.url)
        except Exception as exc:
            return FetchResult(
                document=None,
                error=f"{type(exc).__name__}: {exc}",
            )
        if response.status_code >= 400:
            return FetchResult(document=None, error=f"HTTP {response.status_code} for {item.url}")

        return FetchResult(
            document=RawDocument(
                item=item,
                payload=response.content,
                mime_type="application/octet-stream",
                retrieved_at=utcnow(),
                http_status=response.status_code,
                headers={},
            )
        )

    def health_check(self) -> HealthCheckResult:
        """Probe the file with a range request rather than downloading it."""
        started = utcnow()
        try:
            response = self.http.get(self.download_url, headers={"Range": "bytes=0-1023"})
        except Exception as exc:
            return HealthCheckResult(
                healthy=False, checked_at=started, message=f"{type(exc).__name__}: {exc}"
            )
        return HealthCheckResult(
            healthy=response.status_code < 400,
            checked_at=started,
            latency_ms=response.elapsed_ms,
            http_status=response.status_code,
        )

    def parse(self, document: RawDocument) -> ParseResult:
        """Extract measurement rows from the bulk file."""
        try:
            rows = self.rows_from(document.payload)
        except Exception as exc:
            return ParseResult(document=None, error=f"{type(exc).__name__}: {exc}")
        if not rows:
            return ParseResult(document=None, error="No area rows parsed from the published file")

        return ParseResult(
            document=ParsedDocument(
                raw=document,
                document_type=self.document_type,
                records=rows,
                field_signature=short_hash(",".join(sorted(rows[0]))),
            )
        )

    def normalize(self, document: ParsedDocument) -> NormalizationResult:
        """Turn parsed rows into ``area_total`` records.

        No evidence records are emitted. Evidence in Helios is a claim about a
        *site*, and a county's water total is not one. The provenance that
        matters -- which document version this figure came from -- is carried on
        the ``area_totals`` row itself.
        """
        records = [
            NormalizedRecord(
                entity_type="area_total",
                source_native_id=(
                    f"{row['area_kind']}:{row['area_code']}:{row['metric']}:"
                    f"{row['sector']}:{row['reference_year']}"
                ),
                payload=row,
                fields=[
                    ExtractedField(
                        name="value",
                        value=row["value"],
                        assertion_class=AssertionClass.REPORTED,
                        extraction_method=ExtractionMethod.TABLE_PARSE,
                        confidence=1.0,
                    )
                ],
            )
            for row in document.records
        ]
        return NormalizationResult(records=records)


def _clean_number(value: Any) -> float | None:
    """Coerce an agency's cell value to a float, or None if it is not one."""
    if value is None:
        return None
    if isinstance(value, int | float):
        return float(value)
    text = str(value).strip().replace(",", "")
    # Agencies use several different placeholders for "not applicable".
    if text in {"", "-", "--", "NA", "N/A", "NM", "W", "*"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


# ------------------------------------------------------------------- water --

USGS_WATER_URL = (
    "https://www.sciencebase.gov/catalog/file/get/5af3311be4b0da30c1b245d8"
    "?f=__disk__eb%2F74%2Feb%2Feb74ebb41169c76aaf374990bd5a71cac82604c1"
)
"""``usco2015v2.0.csv`` from the USGS data release."""

USGS_WATER_YEAR = 2015
"""The most recent USGS compilation published at county resolution. The 2020
national circular did not repeat the county-level breakdown, so this is the
newest county figure that exists, not the newest Helios bothered to fetch."""

WATER_CATEGORIES: dict[str, str] = {
    "PS-Wtotl": "public_supply_water_withdrawal",
    "IN-Wtotl": "industrial_water_withdrawal",
    "PT-Wtotl": "thermoelectric_water_withdrawal",
    "TO-Wtotl": "total_water_withdrawal",
}
"""USGS column -> Helios metric. Public supply is the one a data centre on
municipal water actually competes for; the others give it context."""


class UsgsCountyWaterConnector(BulkAreaTotalsConnector):
    """Reads county-level water withdrawals from the USGS 2015 data release."""

    def __init__(self, *, counties: tuple[str, ...] | None = None, **kwargs: Any) -> None:
        """Initialise the connector.

        Args:
            counties: County FIPS codes to keep. ``None`` keeps every county,
                which is 3,223 rows and how a national build would run.
            **kwargs: Passed to :class:`BaseConnector`.
        """
        super().__init__(**kwargs)
        self.counties = counties

    @property
    def download_url(self) -> str:
        """The USGS ScienceBase file."""
        return USGS_WATER_URL

    @property
    def reference_year(self) -> int:
        """2015, and there is no newer county-level figure."""
        return USGS_WATER_YEAR

    def get_metadata(self) -> Any:
        """Return the connector description."""
        from helios_common.vocabulary import AccessMethod, ConnectorStatus, SourceCategory
        from helios_connectors.types import ConnectorMetadata

        return ConnectorMetadata(
            slug="usgs-county-water-use",
            source_slug="usgs-county-water-use",
            name="USGS Estimated Use of Water, County-Level",
            agency="United States Geological Survey",
            jurisdiction="United States",
            category=SourceCategory.WATER,
            access_method=AccessMethod.BULK_DOWNLOAD,
            base_url=USGS_WATER_URL,
            connector_version="0.1.0",
            parser_version="0.1.0",
            status=ConnectorStatus.IMPLEMENTED,
            update_frequency="every five years, historically",
            rate_limit_per_second=0.5,
            license_name="US Government public domain",
            license_url=(
                "https://www.usgs.gov/information-policies-and-instructions/copyrights-and-credits"
            ),
            attribution_required=True,
            attribution_text=(
                "Dieter, C.A., and others, 2018, Estimated use of water in the United "
                "States county-level data for 2015 (ver. 2.0): U.S. Geological Survey "
                "data release, https://doi.org/10.5066/F7TB15V5."
            ),
            robots_policy_status="allowed",
            geographic_coverage="All 3,223 US counties and county equivalents.",
            historical_coverage=(
                "Five-yearly compilations; 2015 is the most recent county-level release."
            ),
            reliability_score=0.9,
            known_schema_issues=(
                "The CSV carries a citation line above the real header row. Withdrawals "
                "are in million gallons per day and population in thousands. The 2020 "
                "compilation did not repeat the county breakdown, so 2015 is the newest "
                "county figure that exists."
            ),
        )

    def rows_from(self, payload: bytes) -> list[dict[str, Any]]:
        """Read the USGS county CSV into measurement rows."""
        text = payload.decode("utf-8-sig", errors="replace")
        handle = io.StringIO(text)
        # The first line is a citation, not a header. Consuming it blind would be
        # fragile, so confirm the real header follows before trusting the parse.
        first = handle.readline()
        if "STATE" in first.split(",")[0].upper():
            handle.seek(0)
        reader = csv.DictReader(handle)

        rows: list[dict[str, Any]] = []
        for record in reader:
            fips = (record.get("FIPS") or "").strip()
            if not fips:
                continue
            if self.counties is not None and fips not in self.counties:
                continue

            county = (record.get("COUNTY") or "").strip()
            state = (record.get("STATE") or "").strip()
            area_name = f"{county}, {state}" if county and state else fips

            for column, metric in WATER_CATEGORIES.items():
                value = _clean_number(record.get(column))
                if value is None:
                    continue
                rows.append(
                    {
                        "area_kind": "county",
                        "area_code": fips,
                        "area_name": area_name,
                        "metric": metric,
                        "sector": "all",
                        "value": value,
                        "unit": "Mgal/d",
                        "reference_year": USGS_WATER_YEAR,
                    }
                )

            population = _clean_number(record.get("TP-TotPop"))
            if population is not None:
                rows.append(
                    {
                        "area_kind": "county",
                        "area_code": fips,
                        "area_name": area_name,
                        "metric": "population",
                        "sector": "all",
                        # Published in thousands; stored as people so no consumer
                        # has to remember the multiplier.
                        "value": population * 1000.0,
                        "unit": "people",
                        "reference_year": USGS_WATER_YEAR,
                    }
                )
        return rows


# ------------------------------------------------------------- electricity --

EIA_SALES_URL = "https://www.eia.gov/electricity/data/state/sales_annual.xlsx"
"""Retail sales to ultimate customers, by state, sector and provider."""

EIA_SECTORS = ("residential", "commercial", "industrial", "transportation", "total")


class EiaStateElectricityConnector(BulkAreaTotalsConnector):
    """Reads state-level retail electricity sales from EIA."""

    def __init__(
        self,
        *,
        states: tuple[str, ...] | None = None,
        year: int | None = None,
        **kwargs: Any,
    ) -> None:
        """Initialise the connector.

        Args:
            states: Two-letter state codes to keep. ``None`` keeps every state.
            year: Reference year to keep. ``None`` keeps the latest in the file.
            **kwargs: Passed to :class:`BaseConnector`.
        """
        super().__init__(**kwargs)
        self.states = states
        self._year = year
        self._resolved_year: int | None = None

    @property
    def download_url(self) -> str:
        """The EIA sales workbook."""
        return EIA_SALES_URL

    @property
    def reference_year(self) -> int:
        """The year kept from the workbook.

        Falls back to the requested year before the file has been read, since
        discovery names the document before parsing has happened.
        """
        return self._resolved_year or self._year or 0

    def get_metadata(self) -> Any:
        """Return the connector description."""
        from helios_common.vocabulary import AccessMethod, ConnectorStatus, SourceCategory
        from helios_connectors.types import ConnectorMetadata

        return ConnectorMetadata(
            slug="eia-state-electricity-sales",
            source_slug="eia-state-electricity-sales",
            name="EIA Retail Electricity Sales by State",
            agency="United States Energy Information Administration",
            jurisdiction="United States",
            category=SourceCategory.INFRASTRUCTURE_REFERENCE,
            access_method=AccessMethod.BULK_DOWNLOAD,
            base_url=EIA_SALES_URL,
            connector_version="0.1.0",
            parser_version="0.1.0",
            status=ConnectorStatus.IMPLEMENTED,
            update_frequency="annual",
            rate_limit_per_second=0.5,
            license_name="US Government public domain",
            license_url="https://www.eia.gov/about/copyrights_reuse.php",
            robots_policy_status="allowed",
            geographic_coverage="All US states and DC. State resolution only.",
            historical_coverage="1990 to the most recent published year.",
            reliability_score=0.95,
            known_schema_issues=(
                "Published only as xlsx; EIA offers no CSV equivalent. Sales are "
                "reported per state, and no public source breaks retail sales to "
                "county nationally, so this cannot be narrowed to a metro area."
            ),
        )

    def rows_from(self, payload: bytes) -> list[dict[str, Any]]:
        """Read the EIA sales workbook into measurement rows."""
        import openpyxl

        if not zipfile.is_zipfile(io.BytesIO(payload)):
            raise ValueError("EIA sales payload is not an xlsx workbook")

        workbook = openpyxl.load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]

        header: list[str] | None = None
        records: list[dict[str, Any]] = []
        for raw in sheet.iter_rows(values_only=True):
            cells = ["" if c is None else str(c).strip() for c in raw]
            if header is None:
                # The title occupies the first row; the header is the row that
                # actually names Year and State.
                if "Year" in cells and "State" in cells:
                    header = cells
                continue
            if not any(cells):
                continue
            records.append(dict(zip(header, raw, strict=False)))
        workbook.close()

        if header is None:
            raise ValueError("Could not locate the Year/State header row in the EIA workbook")

        # "Total Electric Industry" is the all-providers roll-up; the other
        # provider categories are subsets of it and would double count.
        totals = [
            r
            for r in records
            if str(r.get("Industry Sector Category") or "").strip() == "Total Electric Industry"
        ]
        years = {int(y) for r in totals if (y := _clean_number(r.get("Year"))) is not None}
        if not years:
            raise ValueError("No usable years in the EIA workbook")
        target = self._year if self._year in years else max(years)
        self._resolved_year = target

        rows: list[dict[str, Any]] = []
        for record in totals:
            year = _clean_number(record.get("Year"))
            if year is None or int(year) != target:
                continue
            state = str(record.get("State") or "").strip().upper()
            if not state or state == "US":
                continue
            if self.states is not None and state not in self.states:
                continue

            for sector in EIA_SECTORS:
                value = _clean_number(record.get(sector.capitalize()))
                if value is None:
                    continue
                rows.append(
                    {
                        "area_kind": "state",
                        "area_code": state,
                        "area_name": state,
                        "metric": "electricity_retail_sales",
                        "sector": sector,
                        "value": value,
                        "unit": "MWh/yr",
                        "reference_year": target,
                    }
                )
        return rows
